import os
import re
import tempfile
import traceback
from datetime import datetime
from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

ALLOWED_EXTENSIONS = {'pdf', 'xls', 'xlsx', 'csv'}

# Transaction categorization patterns for SBI statements
CATEGORY_PATTERNS = {
    # Deposits - ordered from most specific to least specific
    'Salary': [r'(?i)salary', r'(?i)SAL\s*TRF', r'(?i)NEFT.*SAL', r'(?i)SAL\s*FOR'],
    'DEP TFR  HRMS Mobile': [r'(?i)HRMS\s*Mobile'],
    'DEP TFR  HRMS Labour': [r'(?i)HRMS\s*Labour'],
    'DEP TFR  HRMS Cleansing': [r'(?i)HRMS\s*Cleansing'],
    'DEP TFR  HRMS Briefcase': [r'(?i)HRMS\s*Briefcase'],
    'DEP TFR  HRMS Furniture': [r'(?i)HRMS\s*Furniture'],
    'DEP TFR  HRMS  Utility': [r'(?i)HRMS.*Utility'],
    'DEP TFR  HRMS Pest': [r'(?i)HRMS.*Pest'],
    'DEP TFR  HRMS': [r'(?i)HRMS(?!\s*\w)', r'(?i)DEP\s*TFR.*PF\s*No.*HRMS$', r'(?i)PF\s*No.*\d+\s*HRMS$'],
    'DEP BANKS PERFORMANCE PLI': [r'(?i)BANKS?\s*PERFORMANCE\s*PLI', r'(?i)PERFORMANCE\s*PLI'],
    'CDS BASED PLI PAID FOR THE FY': [r'(?i)CDS\s*BASED\s*PLI', r'(?i)PLI\s*PAID'],
    'CEMTEX DEP INTER CIRCLE SPORTS': [r'(?i)CEMTEX.*DEP', r'(?i)INTER\s*CIRCLE\s*SPORTS', r'(?i)HALTING\s*ALLOWANCE'],
    
    # Withdrawals - ordered from most specific to least specific
    'Bank INTEREST': [r'(?i)TO\s*INTEREST', r'(?i)INTEREST\s*(?:DR|DEBIT)?$', r'(?i)INT\s*(?:CHARGE|DR)'],
    'DIRECT DR': [r'(?i)DIRECT\s*DR', r'(?i)SI\s*DR', r'(?i)ECS\s*DR', r'(?i)NACH\s*DR', r'(?i)OFFICER\s*LEVY'],
    'Transfer to own A/c': [r'(?i)TRF\s*TO\s*(?:OWN|SELF)', r'(?i)SELF\s*TRF', r'(?i)OWN\s*A/?C', r'(?i)INB\s*MBS'],
    'WDL TFR': [r'(?i)WDL\s*TFR', r'(?i)NBT\s*TFR', r'(?i)WEL\s*TFR'],
    'UPI Payment': [r'(?i)UPI[/-]', r'(?i)UPI\s*(?:DR|DEBIT)'],
    'NEFT/RTGS': [r'(?i)NEFT', r'(?i)RTGS', r'(?i)IMPS'],
    'ATM Withdrawal': [r'(?i)ATM', r'(?i)CASH\s*WDL'],
}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_text_from_image_pdf(pdf_path, dpi_scale=3):
    """Extract text from image-based PDF using OCR with higher quality"""
    doc = fitz.open(pdf_path)
    all_text = []
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        
        # Higher resolution for better OCR
        pix = page.get_pixmap(matrix=fitz.Matrix(dpi_scale, dpi_scale))
        img_data = pix.tobytes("png")
        img = Image.open(io.BytesIO(img_data))
        
        # OCR with custom config for better table recognition
        custom_config = r'--oem 3 --psm 6'
        try:
            text = pytesseract.image_to_string(img, lang='eng', config=custom_config)
            all_text.append(text)
        except Exception as e:
            print(f"OCR failed for page {page_num}: {e}")
    
    doc.close()
    return '\n'.join(all_text)


def parse_sbi_transactions(text):
    """Parse SBI bank statement transactions from OCR text"""
    transactions = []
    
    lines = text.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Skip irrelevant lines
        if not line or 'Page no' in line or 'Post Date' in line or 'BROUGHT FORWARD' in line.upper():
            i += 1
            continue
        
        # Match date pattern at start of line
        date_match = re.match(r'^(\d{2}-\d{2}-\d{4})\s+(\d{2}-\d{2}-\d{4})?\s*(.*)', line)
        
        if date_match:
            post_date = date_match.group(1)
            value_date = date_match.group(2) or post_date
            rest = date_match.group(3).strip()
            
            # Collect description and amounts from current and following lines
            description_parts = []
            amounts = []
            
            # Extract amounts from current line
            current_amounts = re.findall(r'([\d,]+\.\d{2})', rest)
            amounts.extend(current_amounts)
            
            # Get description part (before amounts)
            desc = re.sub(r'[\d,]+\.\d{2}.*', '', rest).strip()
            if desc:
                description_parts.append(desc)
            
            # Look at next lines for continuation (description often spans 2 lines)
            j = i + 1
            while j < len(lines) and j <= i + 2:
                next_line = lines[j].strip()
                
                # Stop if we hit another date or page marker
                if re.match(r'^\d{2}-\d{2}-\d{4}', next_line) or 'Page no' in next_line or not next_line:
                    break
                
                # Extract amounts
                next_amounts = re.findall(r'([\d,]+\.\d{2})', next_line)
                amounts.extend(next_amounts)
                
                # Get description (text not looking like amounts)
                next_desc = re.sub(r'[\d,]+\.\d{2}', '', next_line).strip()
                # Filter out garbage OCR
                if next_desc and len(next_desc) > 2 and not next_desc.replace(' ', '').replace(',', '').isdigit():
                    description_parts.append(next_desc)
                
                j += 1
            
            i = j - 1  # Will be incremented at end of loop
            
            # Build final description
            full_desc = ' '.join(description_parts)
            full_desc = re.sub(r'\s+', ' ', full_desc).strip()
            
            # Clean OCR artifacts but preserve key words
            full_desc = re.sub(r'[|]', '', full_desc)
            
            # Determine transaction type
            desc_upper = full_desc.upper()
            is_withdrawal = any(kw in desc_upper for kw in 
                ['WDL', 'WITHDRAWAL', 'TO INTEREST', 'DIRECT DR', 'DEBIT', 'LEVY', 'DR$', 'ATM'])
            
            # Parse amounts - typically: [transaction_amount, balance] or just [balance]
            debit = 0.0
            credit = 0.0
            balance = ""
            
            # Remove duplicate amounts
            unique_amounts = list(dict.fromkeys(amounts))
            
            if len(unique_amounts) >= 2:
                trans_amt = float(unique_amounts[0].replace(',', ''))
                balance = unique_amounts[-1]
                if is_withdrawal:
                    debit = trans_amt
                else:
                    credit = trans_amt
            elif len(unique_amounts) == 1:
                balance = unique_amounts[0]
            
            # Only add if we have meaningful data
            if full_desc and (debit > 0 or credit > 0):
                transactions.append({
                    'Date': post_date,
                    'Value_Date': value_date,
                    'Description': full_desc,
                    'Debit': debit,
                    'Credit': credit,
                    'Balance': balance
                })
        
        i += 1
    
    return transactions


def parse_transactions_from_text(text):
    """Parse transactions from OCR text - wrapper for backward compatibility"""
    return parse_sbi_transactions(text)


def parse_pdf_with_tabula(pdf_path):
    """Try to extract tables using tabula-py as fallback"""
    try:
        import tabula
        dfs = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True, lattice=True)
        if dfs:
            return pd.concat(dfs, ignore_index=True)
    except Exception as e:
        print(f"Tabula extraction failed: {e}")
    return None


def extract_transactions_from_pdf(pdf_path):
    """Main PDF extraction function with multiple methods"""
    doc = fitz.open(pdf_path)
    
    # Check if PDF has text layer
    first_page = doc[0]
    text = first_page.get_text()
    
    transactions = []
    
    if text.strip():
        # Text-based PDF - parse text
        all_text = ""
        for page in doc:
            all_text += page.get_text() + "\n"
        transactions = parse_transactions_from_text(all_text)
    else:
        # Image-based PDF - use OCR
        print("Using OCR for image-based PDF...")
        ocr_text = extract_text_from_image_pdf(pdf_path)
        transactions = parse_transactions_from_text(ocr_text)
    
    doc.close()
    
    # If parsing failed, try tabula
    if not transactions:
        print("Trying tabula extraction...")
        df = parse_pdf_with_tabula(pdf_path)
        if df is not None and not df.empty:
            return df
    
    if transactions:
        return pd.DataFrame(transactions)
    
    return pd.DataFrame()


def extract_transactions_from_excel(file_path):
    """Extract transactions from Excel file"""
    try:
        # Try reading with pandas
        xl = pd.ExcelFile(file_path)
        
        # Look for transaction data in sheets
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Check if this looks like a transaction sheet
            cols_lower = [str(c).lower() for c in df.columns]
            if any('date' in c or 'debit' in c or 'credit' in c or 'balance' in c for c in cols_lower):
                return df
        
        # If no matching sheet found, return first sheet
        return pd.read_excel(file_path)
    except Exception as e:
        print(f"Excel reading failed: {e}")
        return pd.DataFrame()


def normalize_columns(df):
    """Normalize column names to standard format"""
    col_mapping = {}
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if 'post' in col_lower and 'date' in col_lower:
            col_mapping[col] = 'Post_Date'
        elif 'value' in col_lower and 'date' in col_lower:
            col_mapping[col] = 'Value_Date'
        elif col_lower in ['date', 'txn date', 'transaction date']:
            col_mapping[col] = 'Date'
        elif col_lower in ['description', 'narration', 'particulars', 'remarks']:
            col_mapping[col] = 'Description'
        elif col_lower in ['debit', 'debit amount', 'withdrawal', 'dr']:
            col_mapping[col] = 'Debit'
        elif col_lower in ['credit', 'credit amount', 'deposit', 'cr']:
            col_mapping[col] = 'Credit'
        elif col_lower in ['balance', 'closing balance', 'running balance']:
            col_mapping[col] = 'Balance'
        elif 'cheque' in col_lower or 'ref' in col_lower:
            col_mapping[col] = 'Reference'
    
    df = df.rename(columns=col_mapping)
    return df


def clean_amount(value):
    """Clean amount string to float"""
    if pd.isna(value) or value == '' or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Remove commas and convert
    try:
        return float(str(value).replace(',', '').replace(' ', ''))
    except:
        return 0.0


def categorize_transaction(description, is_withdrawal=False):
    """Categorize transaction based on description patterns"""
    if pd.isna(description):
        return 'Other Withdrawal' if is_withdrawal else 'Other Deposit'
    
    desc = str(description).upper()
    
    for category, patterns in CATEGORY_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, desc, re.IGNORECASE):
                return category
    
    return 'Other Withdrawal' if is_withdrawal else 'Other Deposit'


def determine_transaction_type(row):
    """Determine if transaction is deposit or withdrawal"""
    debit = clean_amount(row.get('Debit', 0))
    credit = clean_amount(row.get('Credit', 0))
    
    if credit > 0:
        return 'Deposit', credit
    elif debit > 0:
        return 'Withdrawal', debit
    else:
        # Try to infer from description
        desc = str(row.get('Description', '')).upper()
        if any(kw in desc for kw in ['WDL', 'WITHDRAWAL', 'DEBIT', 'DR', 'TRANSFER OUT']):
            return 'Withdrawal', 0
        else:
            return 'Deposit', 0


def process_transactions(df):
    """Process and categorize all transactions"""
    df = normalize_columns(df)
    
    processed = []
    
    for idx, row in df.iterrows():
        # Skip header rows or empty rows
        if pd.isna(row.get('Description')) and pd.isna(row.get('Date')):
            continue
        
        trans_type, amount = determine_transaction_type(row)
        is_withdrawal = trans_type == 'Withdrawal'
        category = categorize_transaction(row.get('Description', ''), is_withdrawal)
        
        # Parse date
        date_val = row.get('Date') or row.get('Post_Date') or row.get('Value_Date')
        if pd.notna(date_val):
            try:
                if isinstance(date_val, str):
                    for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%b-%Y']:
                        try:
                            date_val = datetime.strptime(date_val, fmt)
                            break
                        except:
                            continue
            except:
                pass
        
        processed.append({
            'Date': date_val,
            'Description': row.get('Description', ''),
            'Category': category,
            'Type': trans_type,
            'Amount': amount if amount else clean_amount(row.get('Debit', 0)) or clean_amount(row.get('Credit', 0)),
            'Debit': clean_amount(row.get('Debit', 0)),
            'Credit': clean_amount(row.get('Credit', 0)),
            'Balance': row.get('Balance', '')
        })
    
    return pd.DataFrame(processed)


def create_styled_workbook(processed_df, account_name='', account_no='', bank_name='SBI'):
    """Create styled Excel workbook with deposits, withdrawals, and summary sheets"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    currency_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Separate deposits and withdrawals
    deposits = processed_df[processed_df['Type'] == 'Deposit'].copy()
    withdrawals = processed_df[processed_df['Type'] == 'Withdrawal'].copy()
    
    def create_category_sheet(ws, transactions_df, sheet_title):
        """Create a sheet with transactions grouped by category"""
        # Header info
        ws['A1'] = 'Name'
        ws['B1'] = account_name
        ws['A2'] = 'Bank'
        ws['B2'] = bank_name
        ws['A3'] = 'Account NO'
        ws['B3'] = account_no
        
        # Group by category
        categories = transactions_df.groupby('Category')
        
        col_offset = 2  # Start from column C (index 3)
        
        for category, group in categories:
            # Category header
            ws.cell(row=4, column=col_offset, value=category).font = Font(bold=True)
            
            # Column headers
            ws.cell(row=5, column=col_offset, value='Date').font = header_font
            ws.cell(row=5, column=col_offset).fill = header_fill
            ws.cell(row=5, column=col_offset + 1, value='Amount').font = header_font
            ws.cell(row=5, column=col_offset + 1).fill = header_fill
            
            # Data
            row_num = 6
            total = 0
            for _, trans in group.iterrows():
                ws.cell(row=row_num, column=col_offset, value=trans['Date'])
                ws.cell(row=row_num, column=col_offset + 1, value=trans['Amount'])
                ws.cell(row=row_num, column=col_offset + 1).number_format = currency_format
                total += trans['Amount']
                row_num += 1
            
            # Total row
            ws.cell(row=row_num, column=col_offset, value='Total').font = Font(bold=True)
            ws.cell(row=row_num, column=col_offset + 1, value=total).font = Font(bold=True)
            ws.cell(row=row_num, column=col_offset + 1).number_format = currency_format
            
            col_offset += 3  # Move to next category column group
        
        # Adjust column widths
        for col in range(1, col_offset + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else f'A{chr(64 + col - 26)}'[0]].width = 15
    
    # Create Deposits sheet
    ws_deposits = wb.active
    ws_deposits.title = 'Deposits'
    create_category_sheet(ws_deposits, deposits, 'Deposits')
    
    # Create Withdrawals sheet
    ws_withdrawals = wb.create_sheet('withdrawals')
    create_category_sheet(ws_withdrawals, withdrawals, 'Withdrawals')
    
    # Create Summary sheet
    ws_summary = wb.create_sheet('summary')
    ws_summary['A1'] = 'Name'
    ws_summary['B1'] = account_name
    ws_summary['A2'] = 'Bank'
    ws_summary['B2'] = bank_name
    ws_summary['A3'] = 'Account NO'
    ws_summary['B3'] = account_no
    
    # Summary headers
    ws_summary['D4'] = 'SL No'
    ws_summary['E4'] = 'Particulars'
    ws_summary['F4'] = 'Deposits'
    ws_summary['G4'] = 'Withdrwals'
    
    for cell in ['D4', 'E4', 'F4', 'G4']:
        ws_summary[cell].font = header_font
        ws_summary[cell].fill = header_fill
    
    # Calculate opening balance (if available)
    row_num = 5
    sl_no = 1
    
    # Opening balance row
    ws_summary[f'D{row_num}'] = sl_no
    ws_summary[f'E{row_num}'] = 'Opening Balance'
    ws_summary[f'F{row_num}'] = ''
    ws_summary[f'G{row_num}'] = ''
    row_num += 1
    sl_no += 1
    
    # Deposit categories
    deposit_totals = deposits.groupby('Category')['Amount'].sum()
    for category, total in deposit_totals.items():
        ws_summary[f'D{row_num}'] = sl_no
        ws_summary[f'E{row_num}'] = category
        ws_summary[f'F{row_num}'] = total
        ws_summary[f'F{row_num}'].number_format = currency_format
        row_num += 1
        sl_no += 1
    
    # Withdrawal categories
    withdrawal_totals = withdrawals.groupby('Category')['Amount'].sum()
    for category, total in withdrawal_totals.items():
        ws_summary[f'D{row_num}'] = sl_no
        ws_summary[f'E{row_num}'] = category
        ws_summary[f'G{row_num}'] = total
        ws_summary[f'G{row_num}'].number_format = currency_format
        row_num += 1
        sl_no += 1
    
    # Totals
    row_num += 1
    ws_summary[f'E{row_num}'] = 'Total'
    ws_summary[f'E{row_num}'].font = Font(bold=True)
    ws_summary[f'F{row_num}'] = deposits['Amount'].sum()
    ws_summary[f'F{row_num}'].number_format = currency_format
    ws_summary[f'F{row_num}'].font = Font(bold=True)
    ws_summary[f'G{row_num}'] = withdrawals['Amount'].sum()
    ws_summary[f'G{row_num}'].number_format = currency_format
    ws_summary[f'G{row_num}'].font = Font(bold=True)
    
    # Adjust column widths
    ws_summary.column_dimensions['D'].width = 8
    ws_summary.column_dimensions['E'].width = 35
    ws_summary.column_dimensions['F'].width = 15
    ws_summary.column_dimensions['G'].width = 15
    
    return wb


def extract_account_info_from_text(text):
    """Extract account info from statement text"""
    account_no = ''
    account_name = ''
    
    # Account number patterns
    acc_match = re.search(r'Account\s*(?:No|Number)[:\s]+(\d+)', text, re.IGNORECASE)
    if acc_match:
        account_no = acc_match.group(1)
    
    # Name patterns
    name_match = re.search(r'(?:Mr\.|Mrs\.|Ms\.?)\s*([A-Z\s]+?)(?:\n|#|Address)', text, re.IGNORECASE)
    if name_match:
        account_name = name_match.group(1).strip()
    
    return account_name, account_no


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Allowed: PDF, XLS, XLSX, CSV'}), 400
    
    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Extract transactions based on file type
        ext = filename.rsplit('.', 1)[1].lower()
        
        if ext == 'pdf':
            df = extract_transactions_from_pdf(filepath)
            # Try to extract account info
            with fitz.open(filepath) as doc:
                text = doc[0].get_text() if doc[0].get_text() else extract_text_from_image_pdf(filepath)
                account_name, account_no = extract_account_info_from_text(text)
        elif ext in ['xls', 'xlsx']:
            df = extract_transactions_from_excel(filepath)
            account_name, account_no = '', ''
        elif ext == 'csv':
            df = pd.read_csv(filepath)
            account_name, account_no = '', ''
        else:
            return jsonify({'error': 'Unsupported file type'}), 400
        
        if df.empty:
            return jsonify({'error': 'No transactions found in file'}), 400
        
        # Process transactions
        processed_df = process_transactions(df)
        
        if processed_df.empty:
            return jsonify({'error': 'Could not parse transactions'}), 400
        
        # Create output workbook
        wb = create_styled_workbook(
            processed_df,
            account_name=account_name,
            account_no=account_no,
            bank_name='SBI'
        )
        
        # Save output
        output_filename = f"Bank_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        wb.save(output_path)
        
        # Clean up input file
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'filename': output_filename,
            'transactions': len(processed_df),
            'deposits': len(processed_df[processed_df['Type'] == 'Deposit']),
            'withdrawals': len(processed_df[processed_df['Type'] == 'Withdrawal'])
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    return jsonify({'error': 'File not found'}), 404


@app.route('/health')
def health():
    return jsonify({'status': 'healthy'}), 200


# Startup log
print("Bank Statement Sorter starting up...")
print(f"Templates folder: {app.template_folder}")


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
